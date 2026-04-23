import json
import re
import io
from datetime import datetime, timezone, timedelta, date
import streamlit as st
import pandas as pd

# ─── Configuração da página ───────────────────────────────────────────────────
st.set_page_config(
    page_title="Trello Extractor",
    layout="wide",
)

# ─── CSS ──────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&family=DM+Sans:wght@300;400;500&display=swap');

/* Reset e base */
html, body, [class*="css"] {
    font-family: 'DM Sans', sans-serif;
}

/* Fundo principal */
.stApp {
    background: #0f1117;
    color: #e8e8e8;
}

/* Header customizado */
.hero-header {
    background: linear-gradient(135deg, #1a1d27 0%, #12141c 100%);
    border: 1px solid #2a2d3a;
    border-radius: 16px;
    padding: 2.5rem 2rem 2rem;
    margin-bottom: 2rem;
    position: relative;
    overflow: hidden;
}
.hero-header::before {
    content: '';
    position: absolute;
    top: -60px; right: -60px;
    width: 200px; height: 200px;
    background: radial-gradient(circle, rgba(99,102,241,0.15) 0%, transparent 70%);
    border-radius: 50%;
}
.hero-header::after {
    content: '';
    position: absolute;
    bottom: -40px; left: 30px;
    width: 120px; height: 120px;
    background: radial-gradient(circle, rgba(236,72,153,0.08) 0%, transparent 70%);
    border-radius: 50%;
}
.hero-title {
    font-family: 'Syne', sans-serif;
    font-size: 2.2rem;
    font-weight: 800;
    color: #f0f0f0;
    margin: 0 0 0.3rem;
    letter-spacing: -0.5px;
}
.hero-title span {
    background: linear-gradient(90deg, #6366f1, #ec4899);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
}
.hero-sub {
    font-size: 0.95rem;
    color: #6b7280;
    margin: 0;
    font-weight: 300;
}

/* Cards de métricas */
.metric-grid {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 1rem;
    margin: 1.5rem 0;
}
.metric-card {
    background: #1a1d27;
    border: 1px solid #2a2d3a;
    border-radius: 12px;
    padding: 1.2rem 1.4rem;
    position: relative;
    overflow: hidden;
    transition: border-color 0.2s;
}
.metric-card:hover { border-color: #6366f1; }
.metric-card::before {
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 3px;
    border-radius: 12px 12px 0 0;
}
.metric-card.total::before  { background: #6366f1; }
.metric-card.done::before   { background: #10b981; }
.metric-card.pending::before{ background: #f59e0b; }
.metric-card.sectors::before{ background: #ec4899; }

.metric-value {
    font-family: 'Syne', sans-serif;
    font-size: 2rem;
    font-weight: 700;
    color: #f0f0f0;
    line-height: 1;
    margin-bottom: 0.3rem;
}
.metric-label {
    font-size: 0.78rem;
    color: #6b7280;
    text-transform: uppercase;
    letter-spacing: 0.08em;
    font-weight: 500;
}

/* Badges de status */
.badge {
    display: inline-block;
    padding: 2px 10px;
    border-radius: 999px;
    font-size: 0.72rem;
    font-weight: 600;
    letter-spacing: 0.05em;
    text-transform: uppercase;
}
.badge-done    { background: #064e3b; color: #34d399; border: 1px solid #065f46; }
.badge-pending { background: #451a03; color: #fbbf24; border: 1px solid #78350f; }
.badge-other   { background: #1e1b4b; color: #a5b4fc; border: 1px solid #312e81; }

/* Upload zone */
.upload-zone {
    border: 2px dashed #2a2d3a;
    border-radius: 14px;
    padding: 2.5rem;
    text-align: center;
    background: #13151f;
    transition: border-color 0.2s;
    margin-bottom: 1.5rem;
}
.upload-zone:hover { border-color: #6366f1; }

/* Streamlit overrides */
div[data-testid="stFileUploader"] > label {
    display: none;
}
div[data-testid="stFileUploader"] section {
    background: #13151f !important;
    border: 2px dashed #2a2d3a !important;
    border-radius: 14px !important;
    padding: 2rem !important;
}
div[data-testid="stFileUploader"] section:hover {
    border-color: #6366f1 !important;
}

/* Date picker container */
div[data-testid="stDateInput"] label {
    color: #9ca3af !important;
    font-size: 0.82rem !important;
    font-weight: 500 !important;
    letter-spacing: 0.04em;
    text-transform: uppercase;
}

/* Dataframe */
div[data-testid="stDataFrame"] {
    border-radius: 12px;
    overflow: hidden;
    border: 1px solid #2a2d3a;
}

/* Expander */
div[data-testid="stExpander"] {
    background: #1a1d27 !important;
    border: 1px solid #2a2d3a !important;
    border-radius: 12px !important;
    margin-bottom: 0.5rem !important;
}
div[data-testid="stExpander"] summary {
    font-family: 'DM Sans', sans-serif;
    font-weight: 500;
    color: #e8e8e8 !important;
    padding: 0.9rem 1.2rem !important;
}

/* Botões */
div[data-testid="stDownloadButton"] button {
    background: linear-gradient(135deg, #6366f1, #8b5cf6) !important;
    color: white !important;
    border: none !important;
    border-radius: 8px !important;
    font-family: 'DM Sans', sans-serif;
    font-weight: 500 !important;
    padding: 0.5rem 1.5rem !important;
    transition: opacity 0.2s !important;
}
div[data-testid="stDownloadButton"] button:hover {
    opacity: 0.85 !important;
}

/* Info box */
.info-box {
    background: #1a1d27;
    border: 1px solid #2a2d3a;
    border-left: 4px solid #6366f1;
    border-radius: 8px;
    padding: 0.9rem 1.2rem;
    font-size: 0.85rem;
    color: #9ca3af;
    margin: 0.5rem 0rem;
}

/* Divider */
.section-divider {
    border: none;
    border-top: 1px solid #1f2130;
    margin: 1.5rem 0;
}

.filter-title {
    font-family: 'Syne', sans-serif;
    font-size: 0.8rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.1em;
    color: #6b7280;
    margin-bottom: 0.8rem;
}

/* Scrollbar */
::-webkit-scrollbar { width: 6px; height: 6px; }
::-webkit-scrollbar-track { background: #0f1117; }
::-webkit-scrollbar-thumb { background: #2a2d3a; border-radius: 3px; }
::-webkit-scrollbar-thumb:hover { background: #6366f1; }
</style>
""", unsafe_allow_html=True)

# ─── Utilitários ──────────────────────────────────────────────────────────────

BRASILIA = timezone(timedelta(hours=-3))
DONE_KEYWORDS = {
    "concluído", "concluida", "concluídos", "concluidas",
    "done", "finalizado", "finalizados", "finished", "completo", "completos",
}

def parse_date(date_str: str | None) -> datetime | None:
    if not date_str:
        return None
    try:
        return datetime.fromisoformat(date_str.replace("Z", "+00:00"))
    except ValueError:
        return None

def format_date(dt: datetime | None) -> str:
    if dt is None:
        return ""
    local = dt.astimezone(BRASILIA)
    return local.strftime("%d/%m/%Y %H:%M")

def format_date_short(dt: datetime | None) -> str:
    if dt is None:
        return ""
    local = dt.astimezone(BRASILIA)
    return local.strftime("%d/%m/%Y")
  
def get_title(title: str, sector: str | None) -> str:
  if sector is not None:
    return title.replace(str(sector + ":"), "")
  else:
    return title

def extract_title(title: str) -> tuple:
    m = re.match(r"^\[(.+?)\]", title)
    if m:
        sector = m.group(1).strip()
        return get_title(title, sector), sector 
      
    m = re.match(r"^([A-ZÀ-Ú][^:\-]{0,28})[\s]*[-:][\s]", title)
    if m:
        sector = m.group(1).strip()
        if len(sector) <= 30:
            return get_title(title, sector), sector

    return get_title(title, None), "—" 
  
def build_observation(status: str, done_date: datetime | None, obs: str) -> str:
    st.write(obs)
  
    if status == "CONCLUÍDO":
        date_str = format_date_short(done_date) if done_date else "data não registrada"
        return f"Concluído no dia {date_str}"
      
    # Não concluído → usa descrição se houver
    return obs if obs else f"Status atual: {status}"  

# FIX
def extract_desc(desc: str, status: str, done_date: datetime | None) -> tuple:
  pos_desc_label = desc.find("[DESCRICAO]:")
  pos_desc_content = (pos_desc_label if pos_desc_label > -1 else 0) + len("[DESCRICAO]:") 
  
  pos_obs_label = desc.find("[OBSERVACAO]:")
  # pos_obs_content = pos_obs_label + len("[OBSERVACAO]:")

  st.write(desc[pos_obs_label:].strip())
  
  description = desc[pos_desc_content:pos_obs_label].strip() or ""
  observation = build_observation(status, done_date, desc[pos_obs_label:].strip()) or ""

  return description, observation
  
def find_done_list_ids(lists: list[dict]) -> set[str]:
    done_ids = set()
    for lst in lists:
        name = lst.get("name", "").lower().strip()
        if name in DONE_KEYWORDS or any(kw in name for kw in DONE_KEYWORDS):
            done_ids.add(lst["id"])
    return done_ids
  
def get_done_date(card: dict, actions: list[dict], done_list_ids: set[str]) -> datetime | None:
    card_id = card["id"]

    # 1. Ação de mover para lista done
    for action in actions:
        if action.get("type") != "updateCard":
            continue
        if action.get("data", {}).get("card", {}).get("id") != card_id:
            continue
        after = action.get("data", {}).get("listAfter", {})
        if after.get("id") in done_list_ids:
            return parse_date(action.get("date"))

    # 2. dueComplete + due
    if card.get("dueComplete") and card.get("due"):
        return parse_date(card["due"])

    # 3. Cartão já está na lista done → dateLastActivity
    if card.get("idList") in done_list_ids:
        return parse_date(card.get("dateLastActivity"))

    return None

def get_status(card: dict, done_list_ids: set[str]) -> str:
    labels = card.get("labels", [])
    if labels:
        name = labels[0].get("name", "").strip()
        return name.upper() if name else "SEM ETIQUETA"
    if card.get("idList") in done_list_ids or card.get("dueComplete"):
        return "CONCLUÍDO"
    return "PENDENTE"

def process_trello_json(data: dict, initial_date: date, final_date: date) -> list[dict]:
    lists = data.get("lists", [])
    cards = data.get("cards", [])
    actions = data.get("actions", [])

    done_list_ids = find_done_list_ids(lists)
    results = []

    for card in cards:
        card_title = card.get("name", "").strip()
        card_description = card.get("desc", "").strip()
        
        done_date = get_done_date(card, actions, done_list_ids)
        status = get_status(card, done_list_ids)

        # Filtro de datas — usa done_date se existir, senão dateLastActivity
        ref_dt = done_date or parse_date(card.get("dateLastActivity"))
        if ref_dt:
            ref_local = ref_dt.astimezone(BRASILIA).date()
            if not (initial_date <= ref_local <= final_date):
                continue

        # Extrai o título e o setor que solicitou o atendimento do título do cartão
        title, sector = extract_title(card_title)
        description, observation = extract_desc(card_description, status, done_date)

        results.append({
            "Data de Finalização": format_date(done_date),
            "Atividade Realizada": title,
            "Descrição da Atividade": description,
            "Setor": sector,
            "Status": status,
            "Observação": observation,
        })

    return results

def to_excel(df: pd.DataFrame) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Atividades")

        ws = writer.sheets["Atividades"]

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        header_fill = PatternFill("solid", fgColor="1E1B4B")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        col_widths = {"A": 20, "B": 40, "C": 45, "D": 18, "E": 14, "F": 42}
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        ws.row_dimensions[1].height = 24

        status_colors = {
            "CONCLUÍDO": "D1FAE5",
            "PENDENTE":  "FEF3C7",
        }
        for row in ws.iter_rows(min_row=2):
            status_cell = row[4]  # coluna E
            fill_color = status_colors.get(status_cell.value, "EDE9FE")
            fill = PatternFill("solid", fgColor=fill_color)
            for cell in row:
                cell.border = border
                cell.fill = fill
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            ws.row_dimensions[status_cell.row].height = 36

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

    return output.getvalue()

def status_badge(status: str) -> str:
    cls = "badge-done" if status == "CONCLUÍDO" else ("badge-pending" if status == "PENDENTE" else "badge-other")
    return f'<span class="badge {cls}">{status}</span>'

# ─── Layout ───────────────────────────────────────────────────────────────────

st.markdown("""
<div class="hero-header">
    <div class="hero-title">Trello <span>Extractor</span></div>
    <p class="hero-sub">Importe o JSON exportado do Trello e gere relatórios filtrados por período</p>
</div>
""", unsafe_allow_html=True)

# Upload
st.markdown("**Arquivo(s) JSON do Trello**")
json_files = st.file_uploader(
    "Selecione o arquivo JSON:",
    accept_multiple_files=True,
    type="json",
    label_visibility="collapsed",
)

if not json_files:
    st.stop()

# Filtros
filter_container = st.container(border=True)
filter_container.markdown('<div><div class="filter-title">🗓 Filtrar por período</div>', unsafe_allow_html=True)
dates_range = filter_container.date_input(
    "Período",
    value=(),
    format="DD/MM/YYYY",
    label_visibility="collapsed",
)
st.markdown("</div>", unsafe_allow_html=True)

if not dates_range or len(dates_range) != 2:
    filter_container.markdown("""
    <div class="info-box">
        ← Selecione a <strong>data inicial</strong> e a <strong>data final</strong> para filtrar as atividades.
    </div>
    """, unsafe_allow_html=True)
    st.stop()

initial_date, final_date = dates_range

# ─── Processamento ────────────────────────────────────────────────────────────

all_activities: list[dict] = []

with st.spinner("Processando arquivos…"):
    for json_file in json_files:
        try:
            data = json.load(json_file)
            activities = process_trello_json(data, initial_date, final_date)
            all_activities.extend(activities)
        except Exception as e:
            st.error(f"Erro ao processar **{json_file.name}**: {e}")

if not all_activities:
    st.warning("Nenhuma atividade encontrada no período selecionado.")
    st.stop()

df = pd.DataFrame(all_activities)

# ─── Métricas ─────────────────────────────────────────────────────────────────

total = len(df)
done = (df["Status"] == "CONCLUÍDO").sum()
pending = (df["Status"] == "PENDENTE").sum()
sectors = df["Setor"].nunique()

st.markdown(f"""
<div class="metric-grid">
    <div class="metric-card total">
        <div class="metric-value">{total}</div>
        <div class="metric-label">Total de atividades</div>
    </div>
    <div class="metric-card done">
        <div class="metric-value">{done}</div>
        <div class="metric-label">Concluídas</div>
    </div>
    <div class="metric-card pending">
        <div class="metric-value">{pending}</div>
        <div class="metric-label">Pendentes</div>
    </div>
    <div class="metric-card sectors">
        <div class="metric-value">{sectors}</div>
        <div class="metric-label">Setores</div>
    </div>
</div>
""", unsafe_allow_html=True)

# ─── Filtros adicionais ────────────────────────────────────────────────────────

col_f1, col_f2 = st.columns([1, 1])

with col_f1:
    sector_options = ["Todos"] + sorted(df["Setor"].unique().tolist())
    selected_sector = st.selectbox("Filtrar por setor", sector_options)

with col_f2:
    status_options = ["Todos"] + sorted(df["Status"].unique().tolist())
    selected_status = st.selectbox("Filtrar por status", status_options)

df_filtered = df.copy()
if selected_sector != "Todos":
    df_filtered = df_filtered[df_filtered["Setor"] == selected_sector]
if selected_status != "Todos":
    df_filtered = df_filtered[df_filtered["Status"] == selected_status]

st.markdown(f"""
<div style="font-size:0.8rem; color:#6b7280; margin:0.5rem 0 1rem;">
    Exibindo <strong style="color:#a5b4fc">{len(df_filtered)}</strong> de {total} atividades
</div>
""", unsafe_allow_html=True)

# ─── Tabela expandida por cartão ──────────────────────────────────────────────

tab_table, tab_cards = st.tabs(["📊 Tabela", "🃏 Cartões"])

with tab_table:
    st.dataframe(
        df_filtered,
        width='stretch',
        hide_index=True,
        column_config={
            "Data de Finalização": st.column_config.TextColumn("📅 Data", width="small"),
            "Atividade Realizada": st.column_config.TextColumn("📌 Atividade", width="large"),
            "Descrição da Atividade": st.column_config.TextColumn("📝 Descrição", width="large"),
            "Setor": st.column_config.TextColumn("🏢 Setor", width="small"),
            "Status": st.column_config.TextColumn("🔖 Status", width="small"),
            "Observação": st.column_config.TextColumn("💬 Observação", width="medium"),
        },
        height=480,
    )

with tab_cards:
    for _, row in df_filtered.iterrows():
        badge = status_badge(row["Status"])
        with st.expander(f"{row['Atividade Realizada']} - {row['Setor']}"):
            c1, c2 = st.columns([1, 2])
            with c1:
                st.markdown(f"**📅 Data de Finalização**  \n{row['Data de Finalização'] or '—'}")
                st.markdown(f"**🏢 Setor**  \n{row['Setor']}")
                st.markdown(f"**🔖 Status**")
                st.markdown(badge, unsafe_allow_html=True)
            with c2:
                st.markdown(f"**📝 Descrição**")
                st.markdown(row["Descrição da Atividade"] or "*Sem descrição*")
                st.markdown(f"**💬 Observação**  \n{row['Observação']}")

# ─── Download ─────────────────────────────────────────────────────────────────

st.markdown("<hr class='section-divider'>", unsafe_allow_html=True)
col_d1, col_d2, _ = st.columns(3)

with col_d1:
    excel_data = to_excel(df_filtered)
    period_str = f"{initial_date.strftime('%d%m%Y')}_a_{final_date.strftime('%d%m%Y')}"
    st.download_button(
        label="⬇️ Baixar Excel",
        data=excel_data,
        file_name=f"trello_atividades_{period_str}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

with col_d2:
    csv_data = df_filtered.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")
    st.download_button(
        label="⬇️ Baixar CSV",
        data=csv_data,
        file_name=f"trello_atividades_{period_str}.csv",
        mime="text/csv",
    )